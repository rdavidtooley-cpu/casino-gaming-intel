#!/usr/bin/env python3
"""
Casino Gaming Equities Dashboard Refresh Script
Usage: python3 refresh_casino_dashboard.py
Dependencies: pip install yfinance openpyxl
"""

import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
import json
import os
from datetime import datetime, timedelta
import traceback
import statistics
import re

# =============================================================================
# CONFIGURATION
# =============================================================================

CASINO_TICKERS = [
    # Casino Majors
    'MGM', 'LVS', 'WYNN', 'CZR',
    # Regional Operators
    'PENN', 'BYD', 'RRR', 'CHDN', 'MCRI', 'GDEN', 'FLL', 'BALY',
    # Online Gaming / Sports Betting
    'DKNG', 'FLUT', 'RSI', 'SRAD',
    # Equipment & Technology
    'LNW', 'ACEL',
    # International
    'MLCO',
]

SECTOR_MAP = {
    'MGM': 'Casino Major',
    'LVS': 'Casino Major',
    'WYNN': 'Casino Major',
    'CZR': 'Casino Major',
    'PENN': 'Regional Operator',
    'BYD': 'Regional Operator',
    'RRR': 'Regional Operator',
    'CHDN': 'Regional Operator',
    'MCRI': 'Regional Operator',
    'GDEN': 'Regional Operator',
    'FLL': 'Regional Operator',
    'BALY': 'Regional Operator',
    'DKNG': 'Online/Sports Betting',
    'FLUT': 'Online/Sports Betting',
    'RSI': 'Online/Sports Betting',
    'SRAD': 'Online/Sports Betting',
    'LNW': 'Equipment/Tech',
    'ACEL': 'Equipment/Tech',
    'MLCO': 'International Operator',
}

GAMING_REVENUE_ESTIMATES = {
    'MGM': 16.2, 'LVS': 11.0, 'WYNN': 6.5, 'CZR': 11.0,
    'PENN': 6.4, 'BYD': 3.9, 'RRR': 2.1, 'CHDN': 2.6,
    'MCRI': 1.4, 'GDEN': 1.0, 'FLL': 0.2, 'BALY': 2.5,
    'DKNG': 4.8, 'FLUT': 13.0, 'RSI': 0.7, 'SRAD': 1.0,
    'LNW': 2.9, 'ACEL': 0.7,
    'MLCO': 4.5,
}

COMPANY_METADATA = {
    'MGM': {'display_name': 'MGM Resorts', 'properties': 'Bellagio, MGM Grand, Mandalay Bay, BetMGM'},
    'LVS': {'display_name': 'Las Vegas Sands', 'properties': 'Venetian Macao, Marina Bay Sands, Londoner Macao'},
    'WYNN': {'display_name': 'Wynn Resorts', 'properties': 'Wynn Las Vegas, Encore, Wynn Macau, Wynn Palace'},
    'CZR': {'display_name': 'Caesars Entertainment', 'properties': 'Caesars Palace, Harrahs, Horseshoe, WSOP'},
    'PENN': {'display_name': 'Penn Entertainment', 'properties': 'Hollywood Casino, ESPN BET, L\'Auberge'},
    'BYD': {'display_name': 'Boyd Gaming', 'properties': 'Orleans, Borgata, IP Biloxi, Treasure Chest'},
    'RRR': {'display_name': 'Red Rock Resorts', 'properties': 'Red Rock Casino, Green Valley Ranch, Palace Station'},
    'CHDN': {'display_name': 'Churchill Downs', 'properties': 'Kentucky Derby, TwinSpires, Exact Sciences Racing'},
    'MCRI': {'display_name': 'Monarch Casino', 'properties': 'Atlantis Casino Reno, Monarch Casino Black Hawk'},
    'GDEN': {'display_name': 'Golden Entertainment', 'properties': 'STRAT, Arizona Charlies, PT\'s Taverns'},
    'FLL': {'display_name': 'Full House Resorts', 'properties': 'Grand Lodge, Stockmans, Chamonix'},
    'BALY': {'display_name': 'Bally\'s Corp', 'properties': 'Bally\'s Atlantic City, Twin River, Dover Downs'},
    'DKNG': {'display_name': 'DraftKings', 'properties': 'DraftKings Sportsbook, DK Casino, DFS'},
    'FLUT': {'display_name': 'Flutter Entertainment', 'properties': 'FanDuel, PokerStars, Betfair, Sky Betting'},
    'RSI': {'display_name': 'Rush Street Interactive', 'properties': 'BetRivers, PlaySugarHouse, RushBet'},
    'SRAD': {'display_name': 'Sportradar', 'properties': 'Sports data, odds, streaming, integrity'},
    'LNW': {'display_name': 'Light & Wonder', 'properties': 'Slot machines, systems, iGaming, SciPlay'},
    'ACEL': {'display_name': 'Accel Entertainment', 'properties': 'Video gaming terminals, IL/NV distributed gaming'},
    'MLCO': {'display_name': 'Melco Resorts', 'properties': 'City of Dreams, Studio City, Altira Macau'},
}

DASHBOARD_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(DASHBOARD_DIR, 'Casino_Gaming_Equities.xlsx')
MARKET_DATA_FILE = os.path.join(DASHBOARD_DIR, 'market_data.json')
COMPANY_DATA_FILE = os.path.join(DASHBOARD_DIR, 'company_data.json')
PRICE_HISTORY_FILE = os.path.join(DASHBOARD_DIR, 'price_history.json')

# Styling
HEADER_FILL = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
NORMAL_FONT = Font(name='Arial', size=10)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# DATA FETCHING
# =============================================================================

def linear_trend(prices):
    n = len(prices)
    if n < 10:
        return None
    x = list(range(n))
    mean_x = sum(x) / n
    mean_y = sum(prices) / n
    numerator = sum((xi - mean_x) * (yi - mean_y) for xi, yi in zip(x, prices))
    denominator = sum((xi - mean_x) ** 2 for xi in x)
    if denominator == 0 or mean_y == 0:
        return None
    slope_per_day = numerator / denominator
    return round(slope_per_day * 252 / mean_y * 100, 1)


def safe_div(value, divisor):
    if value and divisor:
        return value / divisor
    return None


def fetch_ticker_data(ticker):
    try:
        stock = yf.Ticker(ticker)
        info = stock.info

        market_cap_raw = info.get('marketCap', 0)
        ev_raw = info.get('enterpriseValue', 0)

        data = {
            'ticker': ticker,
            'company': info.get('longName', '') or info.get('shortName', ticker),
            'price': info.get('currentPrice') or info.get('regularMarketPrice') or 0,
            'market_cap': market_cap_raw / 1e9 if market_cap_raw else 0,
            'sector': SECTOR_MAP.get(ticker, 'Unknown'),
            'gaming_rev': GAMING_REVENUE_ESTIMATES.get(ticker, 0),

            # Valuation
            'pe_trailing': info.get('trailingPE', None),
            'pe_forward': info.get('forwardPE', None),
            'enterprise_value': ev_raw / 1e9 if ev_raw else None,
            'ev_ebitda': info.get('enterpriseToEbitda', None),
            'ev_revenue': info.get('enterpriseToRevenue', None),
            'price_to_book': info.get('priceToBook', None),
            'eps_trailing': info.get('trailingEps', None),
            'eps_forward': info.get('forwardEps', None),
            'price_to_sales': info.get('priceToSalesTrailing12Months', None),
            'book_value_per_share': info.get('bookValue', None),
            'revenue_per_share': info.get('revenuePerShare', None),

            # Growth
            'revenue_growth': info.get('revenueGrowth', None),
            'earnings_growth': info.get('earningsGrowth', None),
            'quarterly_earnings_growth': info.get('earningsQuarterlyGrowth', None),

            # Profitability
            'profit_margin': info.get('profitMargins', None),
            'roe': info.get('returnOnEquity', None),
            'roa': info.get('returnOnAssets', None),
            'ebitda': info.get('ebitda', 0) / 1e9 if info.get('ebitda') else None,
            'ebitda_margins': info.get('ebitdaMargins', None),
            'gross_margins': info.get('grossMargins', None),
            'operating_margins': info.get('operatingMargins', None),
            'gross_profits': info.get('grossProfits', 0) / 1e9 if info.get('grossProfits') else None,
            'operating_cashflow': info.get('operatingCashflow', 0) / 1e9 if info.get('operatingCashflow') else None,
            'net_income_to_common': info.get('netIncomeToCommon', 0) / 1e9 if info.get('netIncomeToCommon') else None,

            # Financial health
            'free_cash_flow': info.get('freeCashflow', 0) / 1e9 if info.get('freeCashflow') else None,
            'total_debt': info.get('totalDebt', 0) / 1e9 if info.get('totalDebt') else None,
            'total_cash': info.get('totalCash', 0) / 1e9 if info.get('totalCash') else None,
            'debt_to_equity': info.get('debtToEquity', None),
            'current_ratio': info.get('currentRatio', None),
            'quick_ratio': info.get('quickRatio', None),
            'total_cash_per_share': info.get('totalCashPerShare', None),
            'shares_outstanding': info.get('sharesOutstanding', None),
            'float_shares': info.get('floatShares', None),

            # Risk & sentiment
            'beta': info.get('beta', None),
            'shares_short': info.get('sharesShort', None),
            'short_ratio': info.get('shortRatio', None),
            'short_pct_of_float': info.get('shortPercentOfFloat', None),
            'audit_risk': info.get('auditRisk', None),
            'board_risk': info.get('boardRisk', None),
            'compensation_risk': info.get('compensationRisk', None),
            'overall_risk': info.get('overallRisk', None),
            'shareholder_rights_risk': info.get('shareHolderRightsRisk', None),

            # Analyst
            'target_high_price': info.get('targetHighPrice', None),
            'target_low_price': info.get('targetLowPrice', None),
            'target_mean_price': info.get('targetMeanPrice', None),
            'target_median_price': info.get('targetMedianPrice', None),
            'recommendation_mean': info.get('recommendationMean', None),
            'recommendation_key': info.get('recommendationKey', None),
            'number_of_analyst_opinions': info.get('numberOfAnalystOpinions', None),
            'average_analyst_rating': info.get('averageAnalystRating', None),

            # Ownership
            'held_pct_insiders': info.get('heldPercentInsiders', None),
            'held_pct_institutions': info.get('heldPercentInstitutions', None),

            # Dividends
            'dividend_yield': info.get('dividendYield', 0),
            'dividend_rate': info.get('dividendRate', None),
            'payout_ratio': info.get('payoutRatio', None),
            'five_year_avg_dividend_yield': info.get('fiveYearAvgDividendYield', None),
            'trailing_annual_dividend_rate': info.get('trailingAnnualDividendRate', None),
            'trailing_annual_dividend_yield': info.get('trailingAnnualDividendYield', None),

            # Technical / price
            'fifty_two_week_high': info.get('fiftyTwoWeekHigh', 0),
            'fifty_two_week_low': info.get('fiftyTwoWeekLow', 0),
            'fifty_day_average': info.get('fiftyDayAverage', None),
            'two_hundred_day_average': info.get('twoHundredDayAverage', None),
            'average_volume_10d': info.get('averageVolume10days', None),
            'average_volume_3mo': info.get('averageVolume', None),
            'day_high': info.get('dayHigh', None),
            'day_low': info.get('dayLow', None),
            'open_price': info.get('open', None),
            'previous_close': info.get('previousClose', None),
            'regular_market_change': info.get('regularMarketChange', None),
            'regular_market_change_pct': info.get('regularMarketChangePercent', None),

            # Earnings
            'most_recent_quarter': None,
            'earnings_timestamp': None,

            # Company info
            'full_time_employees': info.get('fullTimeEmployees', None),
            'industry_yf': info.get('industry', None),
            'sector_yf': info.get('sector', None),
            'country': info.get('country', None),
            'website': info.get('website', None),
        }

        # Convert Unix timestamps
        for ts_field, ts_key in [('most_recent_quarter', 'mostRecentQuarter'),
                                  ('earnings_timestamp', 'earningsTimestamp')]:
            raw_ts = info.get(ts_key)
            if raw_ts and isinstance(raw_ts, (int, float)):
                try:
                    data[ts_field] = datetime.fromtimestamp(raw_ts).strftime('%Y-%m-%d')
                except (OSError, ValueError):
                    data[ts_field] = None
        ex_div = info.get('exDividendDate')
        if ex_div and isinstance(ex_div, (int, float)):
            try:
                data['ex_dividend_date'] = datetime.fromtimestamp(ex_div).strftime('%Y-%m-%d')
            except (OSError, ValueError):
                data['ex_dividend_date'] = None
        else:
            data['ex_dividend_date'] = None

        # % off 52-week high
        if data['price'] and data['fifty_two_week_high']:
            data['pct_off_high'] = round((data['price'] - data['fifty_two_week_high']) / data['fifty_two_week_high'] * 100, 1)
        else:
            data['pct_off_high'] = None

        # 1Y and YTD changes + linear trend
        hist = stock.history(period='1y')
        if len(hist) > 0:
            price_1y_ago = hist.iloc[0]['Close']
            change_1y = ((data['price'] - price_1y_ago) / price_1y_ago * 100) if price_1y_ago else 0
            data['change_1y'] = round(change_1y, 1)
            data['trend_1y'] = linear_trend(hist['Close'].tolist())
        else:
            data['change_1y'] = 0
            data['trend_1y'] = None

        today = datetime.now()
        ytd_start = datetime(today.year, 1, 1)
        hist_ytd = stock.history(start=ytd_start)
        if len(hist_ytd) > 0:
            price_ytd = hist_ytd.iloc[0]['Close']
            change_ytd = ((data['price'] - price_ytd) / price_ytd * 100) if price_ytd else 0
            data['change_ytd'] = round(change_ytd, 1)
        else:
            data['change_ytd'] = 0

        data['timestamp'] = datetime.now().isoformat()
        return data

    except Exception as e:
        print(f'ERROR fetching {ticker}: {str(e)}')
        return None

def fetch_all_data():
    print('Fetching live data for casino gaming stocks...')
    market_data = {}

    for ticker in CASINO_TICKERS:
        data = fetch_ticker_data(ticker)
        if data:
            market_data[ticker] = data
            print(f'  {ticker}: ${data["price"]:.2f} ({data["change_1y"]:+.1f}%)')
        else:
            print(f'  {ticker}: FAILED')

    return market_data

# =============================================================================
# EXCEL GENERATION
# =============================================================================

def format_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def format_currency(ws, col, start_row, end_row):
    for row in range(start_row, end_row + 1):
        ws.cell(row, col).number_format = '$#,##0.00'
        ws.cell(row, col).border = BORDER
        ws.cell(row, col).font = NORMAL_FONT

def format_percent(ws, col, start_row, end_row):
    for row in range(start_row, end_row + 1):
        ws.cell(row, col).number_format = '0.0%'
        ws.cell(row, col).border = BORDER
        ws.cell(row, col).font = NORMAL_FONT

def fmt_val(value, fallback='N/A'):
    return value if value is not None else fallback


def create_market_overview(wb, market_data):
    ws = wb.active
    ws.title = 'Market Overview'

    headers = [
        'Ticker', 'Company', 'Price ($)', 'Market Cap ($B)', 'EV ($B)',
        '1Y Change (%)', 'YTD Change (%)', '1Y Trend (%/yr)',
        'P/E', 'Fwd P/E', 'EV/EBITDA', 'Beta',
        '52W High', '52W Low', '% Off High',
        'Gaming Rev ($B)', 'Sector',
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    format_header(ws, 1, len(headers))

    row = 2
    for ticker in CASINO_TICKERS:
        if ticker not in market_data:
            continue
        d = market_data[ticker]
        vals = [
            d['ticker'], d['company'], d['price'], d['market_cap'],
            fmt_val(d['enterprise_value']),
            d['change_1y'] / 100, d['change_ytd'] / 100,
            fmt_val(d['trend_1y']),
            fmt_val(d['pe_trailing']), fmt_val(d['pe_forward']),
            fmt_val(d['ev_ebitda']), fmt_val(d['beta']),
            d['fifty_two_week_high'], d['fifty_two_week_low'],
            d['pct_off_high'] / 100 if d['pct_off_high'] is not None else 'N/A',
            d['gaming_rev'], d['sector'],
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row, col, val)
            cell = ws.cell(row, col)
            cell.border = BORDER
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    sr = row
    ws.cell(sr, 1, 'TOTAL/AVG')
    ws.cell(sr, 4, f'=SUM(D2:D{row-1})')
    ws.cell(sr, 5, f'=SUM(E2:E{row-1})')
    ws.cell(sr, 6, f'=AVERAGE(F2:F{row-1})')
    ws.cell(sr, 7, f'=AVERAGE(G2:G{row-1})')
    ws.cell(sr, 9, f'=AVERAGEIF(I2:I{row-1},">0")')
    ws.cell(sr, 10, f'=AVERAGEIF(J2:J{row-1},">0")')
    ws.cell(sr, 11, f'=AVERAGEIF(K2:K{row-1},">0")')
    ws.cell(sr, 16, f'=SUM(P2:P{row-1})')

    for col in range(1, len(headers) + 1):
        cell = ws.cell(sr, col)
        cell.border = BORDER
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.fill = PatternFill(start_color='e8e8e8', end_color='e8e8e8', fill_type='solid')

    format_currency(ws, 3, 2, sr)
    format_currency(ws, 4, 2, sr)
    format_currency(ws, 5, 2, sr)
    format_percent(ws, 6, 2, sr)
    format_percent(ws, 7, 2, sr)
    format_currency(ws, 13, 2, sr)
    format_currency(ws, 14, 2, sr)
    format_percent(ws, 15, 2, sr)
    format_currency(ws, 16, 2, sr)

    widths = [10, 28, 11, 14, 12, 13, 13, 14, 8, 9, 11, 8, 11, 11, 11, 15, 22]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = 'A2'

def create_valuation(wb, market_data):
    ws = wb.create_sheet('Valuation')

    headers = [
        'Ticker', 'Company', 'Price ($)', 'Market Cap ($B)', 'EV ($B)',
        'P/E (TTM)', 'P/E (Fwd)', 'EV/EBITDA', 'EV/Revenue', 'P/B',
        'P/S', 'EPS (TTM)', 'EPS (Fwd)', 'Rev/Share',
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    format_header(ws, 1, len(headers))

    row = 2
    for ticker in CASINO_TICKERS:
        if ticker not in market_data:
            continue
        d = market_data[ticker]
        vals = [
            d['ticker'], d['company'], d['price'], d['market_cap'],
            fmt_val(d['enterprise_value']),
            fmt_val(d['pe_trailing']), fmt_val(d['pe_forward']),
            fmt_val(d['ev_ebitda']), fmt_val(d['ev_revenue']),
            fmt_val(d['price_to_book']),
            fmt_val(d.get('price_to_sales')),
            fmt_val(d.get('eps_trailing')),
            fmt_val(d.get('eps_forward')),
            fmt_val(d.get('revenue_per_share')),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val)
            cell.border = BORDER
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    sr = row
    ws.cell(sr, 1, 'AVERAGE')
    for c in [6, 7, 8, 9, 10, 11]:
        ws.cell(sr, c, f'=AVERAGEIF({get_column_letter(c)}2:{get_column_letter(c)}{row-1},">0")')
    ws.cell(sr, 4, f'=SUM(D2:D{row-1})')
    ws.cell(sr, 5, f'=SUM(E2:E{row-1})')
    for col in range(1, len(headers) + 1):
        cell = ws.cell(sr, col)
        cell.border = BORDER
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.fill = PatternFill(start_color='e8e8e8', end_color='e8e8e8', fill_type='solid')

    format_currency(ws, 3, 2, sr)
    format_currency(ws, 4, 2, sr)
    format_currency(ws, 5, 2, sr)
    format_currency(ws, 12, 2, sr)
    format_currency(ws, 13, 2, sr)
    format_currency(ws, 14, 2, sr)

    widths = [10, 28, 11, 14, 12, 10, 10, 11, 11, 8, 8, 11, 11, 11]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = 'A2'


def create_financial_health(wb, market_data):
    ws = wb.create_sheet('Financial Health')

    headers = [
        'Ticker', 'Company',
        'Rev Growth', 'Earnings Growth', 'Q EPS Growth',
        'Gross Margin', 'EBITDA Margin', 'Op Margin', 'Profit Margin', 'ROE', 'ROA',
        'EBITDA ($B)', 'Op CF ($B)', 'FCF ($B)',
        'Total Debt ($B)', 'Total Cash ($B)', 'Debt/Equity',
        'Current Ratio', 'Quick Ratio',
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    format_header(ws, 1, len(headers))

    row = 2
    for ticker in CASINO_TICKERS:
        if ticker not in market_data:
            continue
        d = market_data[ticker]
        vals = [
            d['ticker'], d['company'],
            fmt_val(d['revenue_growth']),
            fmt_val(d['earnings_growth']),
            fmt_val(d.get('quarterly_earnings_growth')),
            fmt_val(d.get('gross_margins')),
            fmt_val(d.get('ebitda_margins')),
            fmt_val(d.get('operating_margins')),
            fmt_val(d['profit_margin']),
            fmt_val(d['roe']),
            fmt_val(d.get('roa')),
            fmt_val(d.get('ebitda')),
            fmt_val(d.get('operating_cashflow')),
            fmt_val(d['free_cash_flow']),
            fmt_val(d['total_debt']),
            fmt_val(d['total_cash']),
            fmt_val(d['debt_to_equity']),
            fmt_val(d.get('current_ratio')),
            fmt_val(d.get('quick_ratio')),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val)
            cell.border = BORDER
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    for c in [3, 4, 5, 6, 7, 8, 9, 10, 11]:
        for r in range(2, row):
            cell = ws.cell(r, c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.0%'

    for c in [12, 13, 14, 15, 16]:
        format_currency(ws, c, 2, row - 1)

    widths = [10, 28, 11, 13, 12, 12, 13, 11, 12, 8, 8, 12, 11, 11, 14, 13, 12, 12, 11]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = 'A2'


def create_big_movers(wb, market_data):
    ws = wb.create_sheet('Big Movers')

    headers = ['Ticker', 'Company', '1Y Change (%)', '1Y Trend (%/yr)', 'Direction', 'Notes']
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    format_header(ws, 1, len(headers))

    movers = [data for ticker, data in market_data.items() if abs(data['change_1y']) > 20]
    movers_sorted = sorted(movers, key=lambda x: x['change_1y'], reverse=True)

    for row, data in enumerate(movers_sorted, 2):
        ws.cell(row, 1, data['ticker'])
        ws.cell(row, 2, data['company'])
        ws.cell(row, 3, data['change_1y'] / 100)
        ws.cell(row, 4, fmt_val(data['trend_1y']))
        direction = 'Gainer' if data['change_1y'] > 0 else 'Loser'
        ws.cell(row, 5, direction)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            cell.border = BORDER
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='left', vertical='center')

        direction_cell = ws.cell(row, 5)
        if direction == 'Gainer':
            direction_cell.fill = PatternFill(start_color='c6efce', end_color='c6efce', fill_type='solid')
        else:
            direction_cell.fill = PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')

    format_percent(ws, 3, 2, len(movers_sorted) + 1)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 35
    ws.freeze_panes = 'A2'


def create_sector_summary(wb, market_data):
    ws = wb.create_sheet('Sector Summary')

    sectors = {}
    for ticker, data in market_data.items():
        sector = data['sector']
        if sector not in sectors:
            sectors[sector] = []
        sectors[sector].append(data)

    headers = [
        'Sector', '# Companies', 'Avg Mkt Cap ($B)', 'Avg EV ($B)',
        'Avg P/E', 'Avg EV/EBITDA', 'Avg 1Y Change (%)',
        'Total Gaming Rev ($B)',
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    format_header(ws, 1, len(headers))

    def avg_valid(stocks, key):
        vals = [s[key] for s in stocks if s.get(key) is not None]
        return sum(vals) / len(vals) if vals else None

    row = 2
    for sector in sorted(sectors.keys()):
        stocks = sectors[sector]
        ws.cell(row, 1, sector)
        ws.cell(row, 2, len(stocks))
        ws.cell(row, 3, avg_valid(stocks, 'market_cap') or 0)
        ws.cell(row, 4, fmt_val(avg_valid(stocks, 'enterprise_value')))
        ws.cell(row, 5, fmt_val(avg_valid(stocks, 'pe_trailing')))
        ws.cell(row, 6, fmt_val(avg_valid(stocks, 'ev_ebitda')))
        avg_change = sum(s['change_1y'] for s in stocks) / len(stocks)
        ws.cell(row, 7, avg_change / 100)
        ws.cell(row, 8, sum(s['gaming_rev'] for s in stocks))

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            cell.border = BORDER
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    format_currency(ws, 3, 2, row - 1)
    format_currency(ws, 4, 2, row - 1)
    format_percent(ws, 7, 2, row - 1)
    format_currency(ws, 8, 2, row - 1)

    widths = [22, 14, 15, 13, 10, 13, 16, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = 'A2'


def create_summary_stats(wb, market_data):
    ws = wb.create_sheet('Summary Stats')

    n = len(market_data)
    best = max(market_data.values(), key=lambda x: x['change_1y']) if market_data else None
    worst = min(market_data.values(), key=lambda x: x['change_1y']) if market_data else None

    pe_vals = [d['pe_trailing'] for d in market_data.values() if d['pe_trailing']]
    ev_ebitda_vals = [d['ev_ebitda'] for d in market_data.values() if d['ev_ebitda']]
    beta_vals = [d['beta'] for d in market_data.values() if d['beta']]

    stats = [
        ['Metric', 'Value'],
        ['Total Companies', n],
        ['Total Market Cap ($B)', sum(d['market_cap'] for d in market_data.values())],
        ['Total Enterprise Value ($B)', sum(d['enterprise_value'] for d in market_data.values() if d['enterprise_value'])],
        ['Median Market Cap ($B)', statistics.median([d['market_cap'] for d in market_data.values()])],
        ['Total Gaming Revenue ($B)', sum(d['gaming_rev'] for d in market_data.values())],
        ['', ''],
        ['Avg P/E (TTM)', statistics.mean(pe_vals) if pe_vals else 'N/A'],
        ['Median P/E (TTM)', statistics.median(pe_vals) if pe_vals else 'N/A'],
        ['Avg EV/EBITDA', statistics.mean(ev_ebitda_vals) if ev_ebitda_vals else 'N/A'],
        ['Median EV/EBITDA', statistics.median(ev_ebitda_vals) if ev_ebitda_vals else 'N/A'],
        ['Avg Beta', round(statistics.mean(beta_vals), 2) if beta_vals else 'N/A'],
        ['', ''],
        ['Best 1Y Performer', f'{best["company"]} ({best["change_1y"]:+.1f}%)' if best else 'N/A'],
        ['Worst 1Y Performer', f'{worst["company"]} ({worst["change_1y"]:+.1f}%)' if worst else 'N/A'],
    ]

    for row, stat in enumerate(stats, 1):
        for col, val in enumerate(stat, 1):
            ws.cell(row, col, val)

    for row in range(1, len(stats) + 1):
        for col in range(1, 3):
            cell = ws.cell(row, col)
            cell.border = BORDER
            cell.font = NORMAL_FONT if row > 1 else Font(name='Arial', size=10, bold=True)
            if row == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            elif col == 1:
                cell.font = Font(name='Arial', size=10, bold=True)

    for row in [3, 4, 5, 6]:
        ws.cell(row, 2).number_format = '$#,##0.0'
    for row in [8, 9, 10, 11, 12]:
        cell = ws.cell(row, 2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.0'

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 35


def create_price_history_sheet(wb, market_data):
    ws = wb.create_sheet('Price History')

    headers = ['Ticker', 'Price', '1Y Change %', 'YTD Change %', '1Y Trend (%/yr)', '% Off 52W High', 'Beta']
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    format_header(ws, 1, len(headers))

    row = 2
    for ticker in CASINO_TICKERS:
        if ticker not in market_data:
            continue
        d = market_data[ticker]
        ws.cell(row, 1, d['ticker'])
        ws.cell(row, 2, d['price'])
        ws.cell(row, 3, d['change_1y'] / 100)
        ws.cell(row, 4, d['change_ytd'] / 100)
        ws.cell(row, 5, fmt_val(d['trend_1y']))
        ws.cell(row, 6, d['pct_off_high'] / 100 if d['pct_off_high'] is not None else 'N/A')
        ws.cell(row, 7, fmt_val(d['beta']))

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            cell.border = BORDER
            cell.font = NORMAL_FONT

        format_currency(ws, 2, row, row)
        format_percent(ws, 3, row, row)
        format_percent(ws, 4, row, row)
        format_percent(ws, 6, row, row)
        row += 1

    widths = [10, 11, 14, 14, 15, 15, 8]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = 'A2'


def create_analyst_sentiment(wb, market_data):
    ws = wb.create_sheet('Analyst & Sentiment')

    headers = [
        'Ticker', 'Company', 'Price ($)', 'Rec Key', 'Rec Mean', '# Analysts',
        'Target High', 'Target Low', 'Target Mean', 'Target Median',
        'Shares Short', 'Short Ratio', 'Short % Float',
        '% Insiders', '% Institutions',
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    format_header(ws, 1, len(headers))

    row = 2
    for ticker in CASINO_TICKERS:
        if ticker not in market_data:
            continue
        d = market_data[ticker]
        vals = [
            d['ticker'], d['company'], d['price'],
            fmt_val(d.get('recommendation_key')),
            fmt_val(d.get('recommendation_mean')),
            fmt_val(d.get('number_of_analyst_opinions')),
            fmt_val(d.get('target_high_price')),
            fmt_val(d.get('target_low_price')),
            fmt_val(d.get('target_mean_price')),
            fmt_val(d.get('target_median_price')),
            fmt_val(d.get('shares_short')),
            fmt_val(d.get('short_ratio')),
            d.get('short_pct_of_float') if d.get('short_pct_of_float') is not None else 'N/A',
            d.get('held_pct_insiders') if d.get('held_pct_insiders') is not None else 'N/A',
            d.get('held_pct_institutions') if d.get('held_pct_institutions') is not None else 'N/A',
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val)
            cell.border = BORDER
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    format_currency(ws, 3, 2, row - 1)
    format_currency(ws, 7, 2, row - 1)
    format_currency(ws, 8, 2, row - 1)
    format_currency(ws, 9, 2, row - 1)
    format_currency(ws, 10, 2, row - 1)
    for c in [13, 14, 15]:
        for r in range(2, row):
            cell = ws.cell(r, c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.0%'
    for r in range(2, row):
        cell = ws.cell(r, 11)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'

    widths = [10, 28, 11, 10, 10, 10, 12, 12, 12, 13, 14, 11, 12, 11, 13]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = 'A2'


def create_technical(wb, market_data):
    ws = wb.create_sheet('Technical')

    headers = [
        'Ticker', 'Company', 'Price', 'Open', 'Day High', 'Day Low',
        'Prev Close', 'Change', 'Change %',
        '50-Day MA', '200-Day MA',
        'Avg Vol (10d)', 'Avg Vol (3mo)',
        'EPS (TTM)', 'EPS (Fwd)',
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    format_header(ws, 1, len(headers))

    row = 2
    for ticker in CASINO_TICKERS:
        if ticker not in market_data:
            continue
        d = market_data[ticker]
        vals = [
            d['ticker'], d['company'], d['price'],
            fmt_val(d.get('open_price')),
            fmt_val(d.get('day_high')),
            fmt_val(d.get('day_low')),
            fmt_val(d.get('previous_close')),
            fmt_val(d.get('regular_market_change')),
            d.get('regular_market_change_pct') if d.get('regular_market_change_pct') is not None else 'N/A',
            fmt_val(d.get('fifty_day_average')),
            fmt_val(d.get('two_hundred_day_average')),
            fmt_val(d.get('average_volume_10d')),
            fmt_val(d.get('average_volume_3mo')),
            fmt_val(d.get('eps_trailing')),
            fmt_val(d.get('eps_forward')),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val)
            cell.border = BORDER
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    for c in [3, 4, 5, 6, 7, 8, 10, 11, 14, 15]:
        format_currency(ws, c, 2, row - 1)
    for r in range(2, row):
        cell = ws.cell(r, 9)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.00%'
    for c in [12, 13]:
        for r in range(2, row):
            cell = ws.cell(r, c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    widths = [10, 28, 10, 10, 10, 10, 11, 10, 10, 11, 11, 14, 14, 11, 11]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = 'A2'


def generate_excel(market_data):
    wb = Workbook()
    create_market_overview(wb, market_data)
    create_valuation(wb, market_data)
    create_financial_health(wb, market_data)
    create_analyst_sentiment(wb, market_data)
    create_technical(wb, market_data)
    create_big_movers(wb, market_data)
    create_sector_summary(wb, market_data)
    create_summary_stats(wb, market_data)
    create_price_history_sheet(wb, market_data)

    wb.save(OUTPUT_FILE)
    print(f'Generated: {OUTPUT_FILE}')

# =============================================================================
# JSON DATA PERSISTENCE
# =============================================================================

def save_market_data(market_data):
    with open(MARKET_DATA_FILE, 'w') as f:
        json.dump(market_data, f, indent=2)
    print(f'Saved: {MARKET_DATA_FILE}')

def save_company_data(market_data):
    companies = []
    for ticker in CASINO_TICKERS:
        if ticker not in market_data:
            continue
        d = market_data[ticker]
        meta = COMPANY_METADATA.get(ticker, {})
        companies.append({
            'ticker': ticker,
            'company': meta.get('display_name', d['company']),
            'revenue': d['gaming_rev'],
            'country': d.get('country', 'Unknown'),
            'properties': meta.get('properties', ''),
            'employees': d.get('full_time_employees', 0) or 0,
            'sector': d['sector'],
            'market_cap': d['market_cap'],
            'timestamp': d.get('timestamp', ''),
        })
    with open(COMPANY_DATA_FILE, 'w') as f:
        json.dump(companies, f, indent=2)
    print(f'Saved: {COMPANY_DATA_FILE}')

def load_price_history():
    if os.path.exists(PRICE_HISTORY_FILE):
        with open(PRICE_HISTORY_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_price_history(price_history):
    with open(PRICE_HISTORY_FILE, 'w') as f:
        json.dump(price_history, f, indent=2)
    print(f'Updated: {PRICE_HISTORY_FILE}')

def update_price_history(market_data):
    history = load_price_history()
    today = datetime.now().strftime('%Y-%m-%d')

    if today not in history:
        history[today] = {}

    for ticker, data in market_data.items():
        history[today][ticker] = data['price']

    cutoff_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    history_trimmed = {k: v for k, v in history.items() if k >= cutoff_date}

    save_price_history(history_trimmed)

# =============================================================================
# HTML DATA INJECTION (for file:// fallback)
# =============================================================================

def inject_json_into_html(html_path, var_name, json_data):
    if not os.path.exists(html_path):
        print(f'  Skipping (not found): {html_path}')
        return

    with open(html_path, 'r') as f:
        content = f.read()

    pattern = r'(// INJECTED_DATA_START\s*\n)\s*var ' + re.escape(var_name) + r' = .*?;\s*\n(\s*// INJECTED_DATA_END)'
    json_str = json.dumps(json_data, indent=2)
    replacement = r'\1        var ' + var_name + ' = ' + json_str + ';\n        ' + r'\2'

    new_content = re.sub(pattern, replacement, content, flags=re.DOTALL)

    if new_content != content:
        with open(html_path, 'w') as f:
            f.write(new_content)
        print(f'  Injected {var_name} into {os.path.basename(html_path)}')
    else:
        print(f'  No marker found in {os.path.basename(html_path)} — skipping injection')


def inject_data_into_dashboards(market_data):
    print('\nInjecting data into HTML dashboards...')

    equities_html = os.path.join(DASHBOARD_DIR, 'Casino_Equities_Dashboard.html')
    inject_json_into_html(equities_html, 'INJECTED_MARKET_DATA', market_data)

    company_html = os.path.join(DASHBOARD_DIR, 'Casino_Company_Dashboard.html')
    if os.path.exists(COMPANY_DATA_FILE):
        with open(COMPANY_DATA_FILE, 'r') as f:
            company_data = json.load(f)
        inject_json_into_html(company_html, 'INJECTED_COMPANY_DATA', company_data)

    industry_html = os.path.join(DASHBOARD_DIR, 'Casino_Industry_Overview.html')
    industry_json_path = os.path.join(DASHBOARD_DIR, 'industry_data.json')
    if os.path.exists(industry_json_path):
        with open(industry_json_path, 'r') as f:
            industry_data = json.load(f)
        inject_json_into_html(industry_html, 'INJECTED_INDUSTRY_DATA', industry_data)


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    print('=' * 70)
    print('CASINO GAMING EQUITIES DASHBOARD REFRESH')
    print(f'Started: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print('=' * 70)

    try:
        market_data = fetch_all_data()

        if not market_data:
            print('ERROR: No data fetched. Exiting.')
            return

        generate_excel(market_data)
        save_market_data(market_data)
        save_company_data(market_data)
        inject_data_into_dashboards(market_data)
        update_price_history(market_data)

        print('=' * 70)
        print('SUMMARY')
        print(f'Stocks Updated: {len(market_data)}/{len(CASINO_TICKERS)}')
        total_cap = sum(d['market_cap'] for d in market_data.values())
        total_ev = sum(d['enterprise_value'] for d in market_data.values() if d['enterprise_value'])
        print(f'Total Market Cap: ${total_cap:.1f}B')
        print(f'Total Enterprise Value: ${total_ev:.1f}B')
        pe_vals = [d['pe_trailing'] for d in market_data.values() if d['pe_trailing']]
        ev_ebitda_vals = [d['ev_ebitda'] for d in market_data.values() if d['ev_ebitda']]
        if pe_vals:
            print(f'Avg P/E (TTM): {statistics.mean(pe_vals):.1f}x')
        if ev_ebitda_vals:
            print(f'Avg EV/EBITDA: {statistics.mean(ev_ebitda_vals):.1f}x')
        best = max(market_data.values(), key=lambda x: x['change_1y'])
        worst = min(market_data.values(), key=lambda x: x['change_1y'])
        print(f'Best Performer: {best["company"]} ({best["change_1y"]:+.1f}%)')
        print(f'Worst Performer: {worst["company"]} ({worst["change_1y"]:+.1f}%)')
        print(f'Completed: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        print('=' * 70)

    except Exception as e:
        print(f'FATAL ERROR: {str(e)}')
        traceback.print_exc()

if __name__ == '__main__':
    main()
